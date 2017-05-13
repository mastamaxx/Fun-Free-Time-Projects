########################################
# Mark Max 5/12/17
# Lineup Optimizer for Draftkings MLB DFS
# Designed in Pycharm Community Edition
#
# Drawing from an excel model I've created for expected daily MLB results, presorted by adjusted player 
# value, this program will calculate over 2 million lineup combinations and return the optimal 
# lineup for the day.
########################################


import openpyxl

FULL_TEAM = 10
MAX_P = 5
MAX_C = 4
MAX_1B = 5
MAX_2B = 4
MAX_3B = 4
MAX_SS = 4
MAX_OF = 12
MAX_SALARY = 50000


class Player:
    def __init__(self, id, name, score, salary):
        self.id = id
        self.name = name
        self.score = score
        self.salary = salary

    def getID(self):
        return self.id

    def getScore(self):
        return self.score

    def getName(self):
        return self.name

    def getSalary(self):
        return self.salary


def getPitchers(sheet, players):
    for row in range(2, sheet.max_row + 1):
        id = sheet['Q' + str(row)].value
        name = sheet['A' + str(row)].value
        score = sheet['P' + str(row)].value
        salary = sheet['R' + str(row)].value

        players.append(Player(id, name, score, salary))

        if len(players) >= MAX_P:
            break


def getHitters(sheet, catchers, firstbase, secondbase, thirdbase, shortstop, outfield):
    global FULL_TEAM
    global MAX_P
    global MAX_C
    global MAX_1B
    global MAX_2B
    global MAX_3B
    global MAX_SS
    global MAX_OF

    for row in range(2, sheet.max_row + 1):
        id = sheet['T' + str(row)].value
        name = sheet['A' + str(row)].value
        score = sheet['S' + str(row)].value
        salary = sheet['V' + str(row)].value

        if 'C' in sheet['U' + str(row)].value and len(catchers) < MAX_C:
            catchers.append(Player(id, name, score, salary))

        if '1B' in sheet['U' + str(row)].value and len(firstbase) < MAX_1B:
            firstbase.append(Player(id, name, score, salary))

        if '2B' in sheet['U' + str(row)].value and len(secondbase) < MAX_2B:
            secondbase.append(Player(id, name, score, salary))

        if '3B' in sheet['U' + str(row)].value and len(thirdbase) < MAX_3B:
            thirdbase.append(Player(id, name, score, salary))

        if 'SS' in sheet['U' + str(row)].value and len(shortstop) < MAX_SS:
            shortstop.append(Player(id, name, score, salary))

        if 'OF' in sheet['U' + str(row)].value and len(outfield) < MAX_OF:
            outfield.append(Player(id, name, score, salary))

        if (len(catchers) + len(firstbase) + len(secondbase) + len(thirdbase) + len(shortstop) + len(outfield)) \
                >= (MAX_C + MAX_1B + MAX_2B + MAX_3B + MAX_SS + MAX_OF):
            break


file = open("output.txt", 'w')

workbook = openpyxl.load_workbook('MLB Data Model.xlsx')

pitchers = []
sheet = workbook.get_sheet_by_name('P')
getPitchers(sheet, pitchers)

catchers = []
firstbase = []
secondbase = []
shortstops = []
thirdbase = []
outfield = []
sheet = workbook.get_sheet_by_name('H')
getHitters(sheet, catchers, firstbase, secondbase, thirdbase, shortstops, outfield)

numPitchers = len(pitchers)
# numC = len(catchers)
# num1B = len(firstbase)
# num2B = len(secondbase)
# numSS = len(shortstops)
# num3B = len(thirdbase)
numOF = len(outfield)

# print(numPitchers)
# print(numC)
# print(num1B)
# print(num2B)
# print(num3B)
# print(numSS)
# print(numOF)

maxScore = 0
count = 0

for P1 in range(0, numPitchers - 1):
    team = ['', '', '', '', '', '', '', '', '', '']
    team[0] = pitchers[P1]
    for P2 in range(P1 + 1, numPitchers):
        team[1] = pitchers[P2]
        for C in catchers:
            team[2] = C
            for OneB in firstbase:
                if OneB.getID() != C.getID():
                    team[3] = OneB
                else:
                    continue
                for TwoB in secondbase:
                    if TwoB.getID() != OneB.getID() and TwoB.getID() != C.getID():
                        team[4] = TwoB
                    else:
                        continue
                    for SS in shortstops:
                        if SS.getID() != TwoB.getID() and SS.getID() != OneB.getID() and SS.getID() != C.getID():
                            team[5] = SS
                        else:
                            continue
                        for ThreeB in thirdbase:
                            if ThreeB.getID() != SS.getID() and ThreeB.getID() != TwoB.getID() \
                                    and ThreeB.getID() != OneB.getID() and ThreeB.getID() != C.getID():
                                team[6] = ThreeB
                            else:
                                continue
                            for OF1 in range(0, numOF - 2):
                                if outfield[OF1].getID() != ThreeB.getID() and outfield[OF1].getID() != SS.getID()\
                                        and outfield[OF1].getID() != TwoB.getID() \
                                        and outfield[OF1].getID() != OneB.getID() \
                                        and outfield[OF1].getID() != C.getID():
                                    team[7] = outfield[OF1]
                                else:
                                    continue
                                for OF2 in range(OF1 + 1, numOF - 1):
                                    if outfield[OF2].getID() != ThreeB.getID() and outfield[OF2].getID() != SS.getID()\
                                           and outfield[OF2].getID() != TwoB.getID() \
                                            and outfield[OF2].getID() != OneB.getID() \
                                            and outfield[OF2].getID() != C.getID():
                                        team[8] = outfield[OF2]
                                    else:
                                        continue
                                    for OF3 in range(OF2 + 1, numOF):
                                        if outfield[OF3].getID() != ThreeB.getID() \
                                                and outfield[OF3].getID() != SS.getID() \
                                                and outfield[OF3].getID() != TwoB.getID() \
                                                and outfield[OF3].getID() != OneB.getID() \
                                                and outfield[OF3].getID() != C.getID():
                                            team[9] = outfield[OF3]
                                        else:
                                            continue

                                        totalSalary = 0
                                        totalScore = 0

                                        for i in team:
                                            totalSalary += int(i.getSalary())
                                            totalScore += float(i.getScore())

                                        if totalSalary <= MAX_SALARY and totalScore > maxScore:

                                            file.write(
                                                "P: " + str(team[0].getName()) + " " + str(team[0].getScore()) + "\n")
                                            file.write(
                                                "P: " + str(team[1].getName()) + " " + str(team[1].getScore()) + "\n")
                                            file.write(
                                                "C: " + str(team[2].getName()) + " " + str(team[2].getScore()) + "\n")
                                            file.write(
                                                "1B: " + str(team[3].getName()) + " " + str(team[3].getScore()) + "\n")
                                            file.write(
                                                "2B: " + str(team[4].getName()) + " " + str(team[4].getScore()) + "\n")
                                            file.write(
                                                "SS: " + str(team[5].getName()) + " " + str(team[5].getScore()) + "\n")
                                            file.write(
                                                "3B: " + str(team[6].getName()) + " " + str(team[6].getScore()) + "\n")
                                            file.write(
                                                "OF: " + str(team[7].getName()) + " " + str(team[7].getScore()) + "\n")
                                            file.write(
                                                "OF: " + str(team[8].getName()) + " " + str(team[8].getScore()) + "\n")
                                            file.write(
                                                "OF: " + str(team[9].getName()) + " " + str(team[9].getScore()) + "\n")
                                            file.write("Total Score: " + str(totalScore) + "\n")
                                            file.write("Total Salary: " + str(totalSalary) + "\n")
                                            file.write("\n\n\n")

                                            maxScore = totalScore

                                        count += 1


print(count)
