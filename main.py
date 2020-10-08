import openpyxl
import random
import math
import time

GAMES_TO_SIM = 10000
MIN_SALARY_TARGET = 48000
MAX_SALARY_TARGET = 50000
MIN_SCORE_TARGET = 107
SCORE_TARGET_1 = MIN_SCORE_TARGET * 1.1
SCORE_TARGET_2 = MIN_SCORE_TARGET * 1.3
SCORE_TARGET_3 = MIN_SCORE_TARGET * 1.45
SCORE_TARGET_4 = MIN_SCORE_TARGET * 1.6
PLAYERS_PER_DK_TEAM = 10
MAX_PLAYERS_PER_TEAM = 5

PA_INDEX = 8
H_INDEX = 9
SINGLE_INDEX = 10
DOUBLE_INDEX = 11
TRIPLE_INDEX = 12
HR_INDEX = 13
O_INDEX = 14
SO_INDEX = 15
FO_INDEX = 16
GO_INDEX = 17
BB_INDEX = 18
RBI_INDEX = 19
R_INDEX = 20
SB_INDEX = 21
CS_INDEX = 22
B_SCORE_INDEX = 23

P_TBF_INDEX = 6
P_H_INDEX = 7
P_SINGLE_INDEX = 8
P_DOUBLE_INDEX = 9
P_TRIPLE_INDEX = 10
P_HR_INDEX = 11
P_BB_INDEX = 12
P_SO_INDEX = 13
P_O_INDEX = 14
P_ER_INDEX = 15
P_TRAT_INDEX = 16
P_ORAT_INDEX = 17
P_TRAE_INDEX = 18
P_ORAE_INDEX = 19
P_W_INDEX = 20
P_CG_INDEX = 21
P_CGS_INDEX = 22
P_NH_INDEX = 23
P_SCORE_INDEX = 24

SINGLE_CHANCE_INDEX = 7
DOUBLE_CHANCE_INDEX = 8
TRIPLE_CHANCE_INDEX = 9
HR_CHANCE_INDEX = 10
BB_CHANCE_INDEX = 11
SO_CHANCE_INDEX = 12
FO_CHANCE_INDEX = 13
GO_CHANCE_INDEX = 14
SB_CHANCE_INDEX = 15
CS_CHANCE_INDEX = 16
H_DK_ID_INDEX = 17
H_DK_POS_INDEX = 18
H_DK_SALARY_INDEX = 19

PITCHERS_PER_TEAM = 2
SP_INDEX = 1
RP_INDEX = 2

P_TEAMID_INDEX = 0
P_PITCHER_ORDER_INDEX = 1
P_PLAYER_NAME_INDEX = 2
P_TEAM_NAME_INDEX = 3
TBF_TARGET_INDEX = 5
ER_TARGET_INDEX = 6
SINGLE_COEF_INDEX = 7
DOUBLE_COEF_INDEX = 8
TRIPLE_COEF_INDEX = 9
HR_COEF_INDEX = 10
BB_COEF_INDEX = 11
SO_COEF_INDEX = 12
P_DK_ID_INDEX = 13
P_DK_SALARY_INDEX = 14
P_DK_AVG_SCORE_INDEX = 15

POSITION_TEAMID_INDEX = 0
POSITION_BATTING_ORDER_INDEX = 1
POSITION_PLAYER_TEAM_INDEX = 2
POSITION_DK_ID_INDEX = 3
POSITION_DK_SALARY_INDEX = 4
POSITION_DK_AVG_SCORE_INDEX = 5

P1_TEAMID_INDEX = 0
P1_PLAYER_TEAM_INDEX = 1
P2_TEAMID_INDEX = 2
P2_PLAYER_TEAM_INDEX = 3
C_TEAMID_INDEX = 4
C_BATTING_ORDER_INDEX = 5
C_PLAYER_TEAM_INDEX = 6
FB_TEAMID_INDEX = 7
FB_BATTING_ORDER_INDEX = 8
FB_PLAYER_TEAM_INDEX = 9
SB_TEAMID_INDEX = 10
SB_BATTING_ORDER_INDEX = 11
SB_PLAYER_TEAM_INDEX = 12
SS_TEAMID_INDEX = 13
SS_BATTING_ORDER_INDEX = 14
SS_PLAYER_TEAM_INDEX = 15
TB_TEAMID_INDEX = 16
TB_BATTING_ORDER_INDEX = 17
TB_PLAYER_TEAM_INDEX = 18
OF1_TEAMID_INDEX = 19
OF1_BATTING_ORDER_INDEX = 20
OF1_PLAYER_TEAM_INDEX = 21
OF2_TEAMID_INDEX = 22
OF2_BATTING_ORDER_INDEX = 23
OF2_PLAYER_TEAM_INDEX = 24
OF3_TEAMID_INDEX = 25
OF3_BATTING_ORDER_INDEX = 26
OF3_PLAYER_TEAM_INDEX = 27
TOTAL_SALARY_INDEX = 28
SCORE_TARGET_1_COUNT_INDEX = 29
SCORE_TARGET_2_COUNT_INDEX = 30
SCORE_TARGET_3_COUNT_INDEX = 31
SCORE_TARGET_4_COUNT_INDEX = 32
TEAM_TOTAL_SCORE_INDEX = 33


GIDP_CHANCE = .5
FO_GO_RATIO = .5

def getBatters(sheet, batters):

    for row in range(2, sheet.max_row + 1):
        playerid = sheet['A' + str(row)].value
        playername = sheet['B' + str(row)].value
        teamname = sheet['C' + str(row)].value
        gamename = sheet['D' + str(row)].value
        pos = sheet['E' + str(row)].value
        team = sheet['U' + str(row)].value
        battingorder = sheet['V' + str(row)].value
        single = sheet['W' + str(row)].value
        double = sheet['X' + str(row)].value + single
        triple = sheet['Y' + str(row)].value + double
        HR = sheet['Z' + str(row)].value + triple
        BB = sheet['AA' + str(row)].value + HR
        SO = sheet['AB' + str(row)].value + BB
        FO = sheet['AC' + str(row)].value + SO
        GO = sheet['AD' + str(row)].value + FO
        SB = sheet['AE' + str(row)].value * 1.4
        CS = sheet['AF' + str(row)].value * 1.4 + SB
        dk_id = sheet['AG' + str(row)].value
        dk_pos = sheet['AH' + str(row)].value
        dk_salary = sheet['AI' + str(row)].value

        batters.append([playerid, playername, teamname, gamename, pos, team, battingorder, single, double, triple, HR,
                       BB, SO, FO, GO, SB, CS, dk_id, dk_pos, dk_salary])

    return team


def getPitchers(sheet, pitchers):

    for row in range(2, sheet.max_row + 1):

        teamid = sheet['A' + str(row)].value
        pitcherorder = sheet['B' + str(row)].value
        playername = sheet['C' + str(row)].value
        teamname = sheet['D' + str(row)].value
        gamename = sheet['E' + str(row)].value
        tbf = sheet['H' + str(row)].value
        er = sheet['S' + str(row)].value
        single_coef = sheet['U' + str(row)].value
        double_coef = sheet['V' + str(row)].value
        triple_coef = sheet['W' + str(row)].value
        hr_coef = sheet['X' + str(row)].value
        bb_coef = sheet['Y' + str(row)].value
        so_coef = sheet['Z' + str(row)].value
        dk_id = sheet['AB' + str(row)].value
        dk_salary = sheet['AC' + str(row)].value
        dk_avg_score = sheet['R' + str(row)].value

        tbf_target = math.floor(tbf + random.random())
        er_target = math.floor(er + random.random())

        pitchers.append([teamid, pitcherorder, playername, teamname, gamename, tbf_target, er_target, single_coef,
                         double_coef, triple_coef, hr_coef, bb_coef, so_coef, dk_id, dk_salary, dk_avg_score])


def calculateAtBat(gameid, teamid, battingorder, pitcherteamid, pitcherorder):

    global first
    global second
    global third
    global outs
    global batters
    global batterstats
    global pitchers
    global pitcherstats
    global battercount
    global team1runs
    global team2runs

    BatterIndex = getBatterIndex(teamid, battingorder)
    BatterStatsIndex = getStatsIndex(gameid, teamid, battingorder)
    PitcherIndex = getPitcherIndex(pitcherteamid, pitcherorder)
    PitcherStatsIndex = getPitcherStatsIndex(gameid, pitcherteamid, pitcherorder)

    batterstats[BatterStatsIndex][PA_INDEX] = batterstats[BatterStatsIndex][PA_INDEX] + 1
    pitcherstats[PitcherStatsIndex][P_TBF_INDEX] = pitcherstats[PitcherStatsIndex][P_TBF_INDEX] + 1

    rand_num = random.random()

    if rand_num < batters[BatterIndex][SINGLE_CHANCE_INDEX] * pitchers[PitcherIndex][SINGLE_COEF_INDEX]:

        # record the result
        single(BatterStatsIndex, PitcherStatsIndex)

        # move the baserunners
        if third != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            third = [0, 0, 0]

        if second != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(second[0], second[1], second[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            second = [0, 0, 0]

        if first != [0, 0, 0]:
            if outs == 2:
                third = first
            else:
                second = first

        first = [gameid, teamid, battingorder]

        #check for stolen base attempt
        if second == [0, 0, 0]:
            rand_steal = random.random()

            if rand_steal < batters[BatterIndex][SB_CHANCE_INDEX]:
                batterstats[BatterStatsIndex][SB_INDEX] = batterstats[BatterStatsIndex][SB_INDEX] + 1
                second = first
            elif rand_steal < batters[BatterIndex][CS_CHANCE_INDEX]:
                batterstats[BatterStatsIndex][CS_INDEX] = batterstats[BatterStatsIndex][CS_INDEX] + 1
                first = [0, 0, 0]
                outs = outs + 1
                pitcherstats[PitcherStatsIndex][P_O_INDEX] = pitcherstats[PitcherStatsIndex][P_O_INDEX] + 1

    elif rand_num < batters[BatterIndex][DOUBLE_CHANCE_INDEX] * pitchers[PitcherIndex][DOUBLE_COEF_INDEX]:

        # record the result
        double(BatterStatsIndex, PitcherStatsIndex)

        # move the baserunners
        if third != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            third = [0, 0, 0]

        if second != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(second[0], second[1], second[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            second = [0, 0, 0]

        if first != [0, 0, 0]:
            if outs == 2:
                baserunner_scores_w_RBI(getStatsIndex(first[0], first[1], first[2]), BatterStatsIndex, PitcherStatsIndex)
                incrementTeamRuns(teamid)
                first = [0, 0, 0]
            else:
                third = first
                first = [0, 0, 0]

        second = [gameid, teamid, battingorder]

    elif rand_num < batters[BatterIndex][TRIPLE_CHANCE_INDEX] * pitchers[PitcherIndex][TRIPLE_COEF_INDEX]:

        # record the result
        batterstats[BatterStatsIndex][H_INDEX] = batterstats[BatterStatsIndex][H_INDEX] + 1
        batterstats[BatterStatsIndex][TRIPLE_INDEX] = batterstats[BatterStatsIndex][TRIPLE_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_H_INDEX] = pitcherstats[PitcherStatsIndex][P_H_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_TRIPLE_INDEX] = pitcherstats[PitcherStatsIndex][P_TRIPLE_INDEX] + 1

        # move the baserunners
        if third != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            third = [0, 0, 0]

        if second != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(second[0], second[1], second[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            second = [0, 0, 0]

        if first != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(first[0], first[1], first[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            first = [0, 0, 0]

        third = [gameid, teamid, battingorder]

    elif rand_num < batters[BatterIndex][HR_CHANCE_INDEX] * pitchers[PitcherIndex][HR_COEF_INDEX]:

        # record the result
        batterstats[BatterStatsIndex][H_INDEX] = batterstats[BatterStatsIndex][H_INDEX] + 1
        batterstats[BatterStatsIndex][HR_INDEX] = batterstats[BatterStatsIndex][HR_INDEX] + 1
        batterstats[BatterStatsIndex][R_INDEX] = batterstats[BatterStatsIndex][R_INDEX] + 1
        batterstats[BatterStatsIndex][RBI_INDEX] = batterstats[BatterStatsIndex][RBI_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_H_INDEX] = pitcherstats[PitcherStatsIndex][P_H_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_HR_INDEX] = pitcherstats[PitcherStatsIndex][P_HR_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_ER_INDEX] = pitcherstats[PitcherStatsIndex][P_ER_INDEX] + 1
        incrementTeamRuns(teamid)

        # move the baserunners
        if third != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            third = [0, 0, 0]

        if second != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(second[0], second[1], second[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            second = [0, 0, 0]

        if first != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(first[0], first[1], first[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            first = [0, 0, 0]

    elif rand_num < batters[BatterIndex][BB_CHANCE_INDEX] * pitchers[PitcherIndex][BB_COEF_INDEX]:

        # record the result
        batterstats[BatterStatsIndex][BB_INDEX] = batterstats[BatterStatsIndex][BB_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_BB_INDEX] = pitcherstats[PitcherStatsIndex][P_BB_INDEX] + 1

        # move the baserunners
        if third != [0, 0, 0] and second != [0, 0, 0] and third!= [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            third = second
            second = first
            first = [gameid, teamid, battingorder]
        elif second != [0, 0, 0] and first != [0, 0, 0]:
            third = second
            second = first
            first = [gameid, teamid, battingorder]
        else:
            second = first
            first = [gameid, teamid, battingorder]

        # check for stolen base attempt
        if second == [0, 0, 0]:
            rand_steal = random.random()

            if rand_steal < batters[BatterIndex][SB_CHANCE_INDEX]:
                batterstats[BatterStatsIndex][SB_INDEX] = batterstats[BatterStatsIndex][SB_INDEX] + 1
                second = first
            elif rand_steal < batters[BatterIndex][CS_CHANCE_INDEX]:
                firststatsindex = getStatsIndex(first[0], first[1], first[2])
                batterstats[firststatsindex][CS_INDEX] = batterstats[firststatsindex][CS_INDEX] + 1
                first = [0, 0, 0]
                outs = outs + 1
                pitcherstats[PitcherStatsIndex][P_O_INDEX] = batterstats[PitcherStatsIndex][P_O_INDEX] + 1

    elif rand_num < batters[BatterIndex][SO_CHANCE_INDEX] * pitchers[PitcherIndex][SO_COEF_INDEX]:

        batterstats[BatterStatsIndex][O_INDEX] = batterstats[BatterStatsIndex][O_INDEX] + 1
        batterstats[BatterStatsIndex][SO_INDEX] = batterstats[BatterStatsIndex][SO_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_O_INDEX] = pitcherstats[PitcherStatsIndex][P_O_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_SO_INDEX] = pitcherstats[PitcherStatsIndex][P_SO_INDEX] + 1
        outs = outs + 1

    elif rand_num < batters[BatterIndex][SO_CHANCE_INDEX] * pitchers[PitcherIndex][SO_COEF_INDEX] \
            + (1 - batters[BatterIndex][SO_CHANCE_INDEX] * pitchers[PitcherIndex][SO_COEF_INDEX]) * FO_GO_RATIO:

        batterstats[BatterStatsIndex][O_INDEX] = batterstats[BatterStatsIndex][O_INDEX] + 1
        batterstats[BatterStatsIndex][FO_INDEX] = batterstats[BatterStatsIndex][FO_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_O_INDEX] = pitcherstats[PitcherStatsIndex][P_O_INDEX] + 1
        outs = outs + 1

        if outs != 3 and third != [0, 0, 0]:
            baserunner_scores_w_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex, PitcherStatsIndex)
            incrementTeamRuns(teamid)
            third = [0, 0, 0]

    else:

        batterstats[BatterStatsIndex][O_INDEX] = batterstats[BatterStatsIndex][O_INDEX] + 1
        batterstats[BatterStatsIndex][GO_INDEX] = batterstats[BatterStatsIndex][GO_INDEX] + 1
        pitcherstats[PitcherStatsIndex][P_O_INDEX] = pitcherstats[PitcherStatsIndex][P_O_INDEX] + 1
        outs = outs + 1

        # check for double play and/or moving runners
        if outs != 3:
            lead_force_runner = get_lead_force_runner(first, second, third)

            if lead_force_runner == 0:
                # no double play possible so batter is out at first and move any runners
                if third != [0, 0, 0]:
                    baserunner_scores_w_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex,
                                            PitcherStatsIndex)
                    incrementTeamRuns(teamid)
                    third = [0, 0, 0]
                third = second
            else:
                # double play is possible so we will test if it was successful
                rand_gidp = random.random()
                if rand_gidp < GIDP_CHANCE:
                    # gidp happened. add an out and eliminate the lead runner & batter & move any remaining runners
                    outs = outs + 1
                    pitcherstats[PitcherStatsIndex][P_O_INDEX] = pitcherstats[PitcherStatsIndex][P_O_INDEX] + 1
                    if outs != 3:
                        # if this out was the 3rd we don't care what happens next
                        if lead_force_runner == 3:
                            # basses loaded out at home & first - runners move from 2nd->3rd and 1st->2nd
                            third = second
                            second = first
                            first = [0, 0, 0]
                        elif lead_force_runner == 2:
                            # runners on first and 2nd outs at 3rd and 1st. runner on 1st->2nd
                            second = first
                            first = [0, 0, 0]
                        else:
                            #runners could be on 3rd and 1st or just 1st. Outs at 2nd and 1st
                            if third != [0, 0, 0]:
                                baserunner_scores_wo_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex,
                                                         PitcherStatsIndex)
                                incrementTeamRuns(teamid)
                                third = [0, 0, 0]
                                first = [0, 0, 0]
                            else:
                                first = [0, 0, 0]
                else:
                    # gidp didn't happen. lead force runner is out and we will move all other runners + 1 incl batter
                    if lead_force_runner == 3:
                        third = second
                        second = first
                        first = [gameid, teamid, battingorder]
                    elif lead_force_runner == 2:
                        second = first
                        first = [gameid, teamid, battingorder]
                    else:
                        # runners could be on 3rd and 1st or just 1st. Out at 2nd
                        if third != [0, 0, 0]:
                            baserunner_scores_wo_RBI(getStatsIndex(third[0], third[1], third[2]), BatterStatsIndex,
                                                     PitcherStatsIndex)
                            incrementTeamRuns(teamid)
                            third = [0, 0, 0]
                            first = [gameid, teamid, battingorder]
                        else:
                            first = [gameid, teamid, battingorder]


def checkforpitchingchange(gameid, teamid, pitcherorder, teamruns, oppruns):

    global pitcherstats

    if pitcherorder == 1:
        PitcherIndex = getPitcherIndex(teamid, pitcherorder)
        PitcherStatsIndex = getPitcherStatsIndex(gameid, teamid, pitcherorder)
        tbf = pitcherstats[PitcherStatsIndex][P_TBF_INDEX]
        tbf_target = pitchers[PitcherIndex][TBF_TARGET_INDEX] + \
                     3 * (pitchers[PitcherIndex][ER_TARGET_INDEX] - pitcherstats[PitcherStatsIndex][P_ER_INDEX])

        if tbf >= tbf_target:
            pitcherstats[PitcherStatsIndex][P_TRAT_INDEX] = teamruns
            pitcherstats[PitcherStatsIndex][P_ORAT_INDEX] = oppruns
            pitcherorder = 2

    return pitcherorder


def getStatsIndex(gameid, teamid, battingorder):
    return ((gameid - 1) * battercount + (teamid - 1) * 9 + battingorder - 1)


def getBatterIndex(teamid, battingorder):
    return ((teamid - 1) * 9 + battingorder - 1)


def getPitcherStatsIndex(gameid, teamid, pitcherorder):
    return ((gameid - 1) * pitchercount + (teamid - 1) * PITCHERS_PER_TEAM + pitcherorder - 1)


def getPitcherIndex(teamid, pitcherorder):
    return ((teamid - 1) * PITCHERS_PER_TEAM + pitcherorder - 1)


def getTeamStatsIndex(gameid, teamid):
    return ((gameid - 1) * lastteam + teamid - 1)


def single(BatterStatsIndex, PitcherStatsIndex):

    global batterstats
    global pitcherstats

    batterstats[BatterStatsIndex][H_INDEX] = batterstats[BatterStatsIndex][H_INDEX] + 1
    batterstats[BatterStatsIndex][SINGLE_INDEX] = batterstats[BatterStatsIndex][SINGLE_INDEX] + 1
    pitcherstats[PitcherStatsIndex][P_H_INDEX] = pitcherstats[PitcherStatsIndex][P_H_INDEX] + 1
    pitcherstats[PitcherStatsIndex][P_SINGLE_INDEX] = pitcherstats[PitcherStatsIndex][P_SINGLE_INDEX] + 1


def double(BatterStatsIndex, PitcherStatsIndex):
    global batterstats
    global pitcherstats

    batterstats[BatterStatsIndex][H_INDEX] = batterstats[BatterStatsIndex][H_INDEX] + 1
    batterstats[BatterStatsIndex][DOUBLE_INDEX] = batterstats[BatterStatsIndex][DOUBLE_INDEX] + 1
    pitcherstats[PitcherStatsIndex][P_H_INDEX] = pitcherstats[PitcherStatsIndex][P_H_INDEX] + 1
    pitcherstats[PitcherStatsIndex][P_DOUBLE_INDEX] = pitcherstats[PitcherStatsIndex][P_DOUBLE_INDEX] + 1


def baserunner_scores_w_RBI(baserunnerstatsindex, batterstatsindex, pitcherstatsindex):

    global batterstats
    global pitcherstats

    batterstats[baserunnerstatsindex][R_INDEX] = batterstats[baserunnerstatsindex][R_INDEX] + 1
    batterstats[batterstatsindex][RBI_INDEX] = batterstats[batterstatsindex][RBI_INDEX] + 1
    pitcherstats[pitcherstatsindex][P_ER_INDEX] = pitcherstats[pitcherstatsindex][P_ER_INDEX] + 1


def baserunner_scores_wo_RBI(baserunnerstatsindex, batterstatsindex, pitcherstatsindex):

    global batterstats
    global pitcherstats

    batterstats[baserunnerstatsindex][R_INDEX] = batterstats[baserunnerstatsindex][R_INDEX] + 1
    pitcherstats[pitcherstatsindex][P_ER_INDEX] = pitcherstats[pitcherstatsindex][P_ER_INDEX] + 1


def incrementTeamRuns(teamid):

    global team1runs
    global team2runs

    if teamid % 2 == 0:
        team2runs = team2runs + 1
    else:
        team1runs = team1runs + 1


def get_lead_force_runner(first, second, third):

    if third != [0, 0, 0] and second != [0, 0, 0] and first != [0, 0, 0]:
        return 3
    elif second != [0, 0, 0] and first != [0, 0, 0]:
        return 2
    elif first != [0, 0, 0]:
        return 1
    else:
        return 0


def getHitters(sheet, catchers, firstbase, secondbase, shortstop, thirdbase, outfield):

    global max_salary
    global min_salary
    global max_ss_salary
    global min_ss_salary
    global max_tb_salary
    global min_tb_salary
    global max_of_salary
    global min_of_salary
    global max_c_avg_score
    global max_fb_avg_score
    global max_sb_avg_score
    global max_ss_avg_score
    global max_tb_avg_score
    global max_of_avg_score

    for row in range(2, sheet.max_row + 1):


        playername = sheet['B' + str(row)].value
        teamname = sheet['C' + str(row)].value
        teamid = sheet['U' + str(row)].value
        battingorder = sheet['V' + str(row)].value
        dk_id = sheet['AG' + str(row)].value
        dk_pos = sheet['AH' + str(row)].value
        dk_salary = sheet['AI' + str(row)].value
        dk_avg_score = sheet['T' + str(row)].value

        if dk_salary > max_salary:
            max_salary = dk_salary
        if dk_salary < min_salary:
            min_salary = dk_salary

        if 'C' in dk_pos:
            catchers.append([teamid, battingorder, playername + " " + teamname, dk_id, dk_salary, dk_avg_score])

            if dk_avg_score > max_c_avg_score:
                max_c_avg_score = dk_avg_score

        if '1B' in dk_pos:
            firstbase.append([teamid, battingorder, playername + " " + teamname, dk_id, dk_salary, dk_avg_score])

            if dk_avg_score > max_fb_avg_score:
                max_fb_avg_score = dk_avg_score

        if '2B' in dk_pos:
            secondbase.append([teamid, battingorder, playername + " " + teamname, dk_id, dk_salary, dk_avg_score])

            if dk_avg_score > max_sb_avg_score:
                max_sb_avg_score = dk_avg_score

        if 'SS' in dk_pos:
            shortstop.append([teamid, battingorder, playername + " " + teamname, dk_id, dk_salary, dk_avg_score])

            if dk_salary > max_ss_salary:
                max_ss_salary = dk_salary
            if dk_salary < min_ss_salary:
                min_ss_salary = dk_salary

            if dk_avg_score > max_ss_avg_score:
                max_ss_avg_score = dk_avg_score

        if '3B' in dk_pos:
            thirdbase.append([teamid, battingorder, playername + " " + teamname, dk_id, dk_salary, dk_avg_score])

            if dk_salary > max_tb_salary:
                max_tb_salary = dk_salary
            if dk_salary < min_tb_salary:
                min_tb_salary = dk_salary

            if dk_avg_score > max_tb_avg_score:
                max_tb_avg_score = dk_avg_score

        if 'OF' in dk_pos:
            outfield.append([teamid, battingorder, playername + " " + teamname, dk_id, dk_salary, dk_avg_score])

            if dk_salary > max_of_salary:
                max_of_salary = dk_salary
            if dk_salary < min_of_salary:
                min_of_salary = dk_salary

            if dk_avg_score > max_of_avg_score:
                max_of_avg_score = dk_avg_score


def initiateTeams(teams, pitchers, catchers, firstbase, secondbase, shortstop, thirdbase, outfield):

    global max_salary
    global min_salary
    global start_time

    for p1 in range(0, len(pitchers) - 1):
        print("%f minutes - P1: " % ((time.time() - start_time) / 60) + str(p1) + " started")
        if pitchers[p1][P_PITCHER_ORDER_INDEX] == 1:
            p1_id = pitchers[p1][P_DK_ID_INDEX]
            p1_teamid = pitchers[p1][P_TEAMID_INDEX]
            p1_player_team = pitchers[p1][P_PLAYER_NAME_INDEX] + " " + pitchers[p1][P_TEAM_NAME_INDEX]
            p1_salary = pitchers[p1][P_DK_SALARY_INDEX]
            p1_avg_score = pitchers[p1][P_DK_AVG_SCORE_INDEX]
        else:
            continue

        for p2 in range(p1 + 1, len(pitchers)):
            if pitchers[p2][P_PITCHER_ORDER_INDEX] == 1:
                p2_id = pitchers[p2][P_DK_ID_INDEX]
                p2_teamid = pitchers[p2][P_TEAMID_INDEX]
                p2_player_team = pitchers[p2][P_PLAYER_NAME_INDEX] + " " + pitchers[p2][P_TEAM_NAME_INDEX]
                p2_salary = pitchers[p2][P_DK_SALARY_INDEX]
                p2_avg_score = pitchers[p2][P_DK_AVG_SCORE_INDEX]
            else:
                continue

            for catcher in catchers:
                c_id = catcher[POSITION_DK_ID_INDEX]
                c_teamid = catcher[POSITION_TEAMID_INDEX]
                c_batting_order = catcher[POSITION_BATTING_ORDER_INDEX]
                c_player_team = catcher[POSITION_PLAYER_TEAM_INDEX]
                c_salary = catcher[POSITION_DK_SALARY_INDEX]
                c_avg_score = catcher[POSITION_DK_AVG_SCORE_INDEX]

                for fb in firstbase:

                    total_avg_score = p1_avg_score + p2_avg_score + c_avg_score
                    if total_avg_score + max_fb_avg_score + max_sb_avg_score + max_ss_avg_score + max_tb_avg_score \
                            + max_of_avg_score * 3 < MIN_SCORE_TARGET:
                        continue

                    fb_id = fb[POSITION_DK_ID_INDEX]
                    if c_id != fb_id:
                        fb_teamid = fb[POSITION_TEAMID_INDEX]
                        fb_batting_order = fb[POSITION_BATTING_ORDER_INDEX]
                        fb_player_team = fb[POSITION_PLAYER_TEAM_INDEX]
                        fb_salary = fb[POSITION_DK_SALARY_INDEX]
                        fb_avg_score = fb[POSITION_DK_AVG_SCORE_INDEX]
                    else:
                        continue

                    for sb in secondbase:

                        total_avg_score = p1_avg_score + p2_avg_score + c_avg_score + fb_avg_score
                        if total_avg_score + max_sb_avg_score + max_ss_avg_score + max_tb_avg_score \
                                + max_of_avg_score * 3 < MIN_SCORE_TARGET:
                            continue

                        sb_id = sb[POSITION_DK_ID_INDEX]
                        if c_id != sb_id and fb_id != sb_id:
                            sb_teamid = sb[POSITION_TEAMID_INDEX]
                            sb_batting_order = sb[POSITION_BATTING_ORDER_INDEX]
                            sb_player_team = sb[POSITION_PLAYER_TEAM_INDEX]
                            sb_salary = sb[POSITION_DK_SALARY_INDEX]
                            sb_avg_score = sb[POSITION_DK_AVG_SCORE_INDEX]
                        else:
                            continue

                        for ss in shortstop:

                            total_avg_score = p1_avg_score + p2_avg_score + c_avg_score + fb_avg_score + sb_avg_score
                            if total_avg_score + max_ss_avg_score + max_tb_avg_score + max_of_avg_score * 3 \
                                    < MIN_SCORE_TARGET:
                                continue

                            total_salary = p1_salary + p2_salary + c_salary + fb_salary + sb_salary
                            if total_salary + max_ss_salary + max_tb_salary + max_of_salary * 3 < MIN_SALARY_TARGET \
                                    or total_salary + min_ss_salary + min_tb_salary + min_of_salary * 3 \
                                    > MAX_SALARY_TARGET:
                                continue

                            ss_id = ss[POSITION_DK_ID_INDEX]
                            if c_id != ss_id and fb_id != ss_id and sb_id != ss_id:
                                ss_teamid = ss[POSITION_TEAMID_INDEX]
                                ss_batting_order = ss[POSITION_BATTING_ORDER_INDEX]
                                ss_player_team = ss[POSITION_PLAYER_TEAM_INDEX]
                                ss_salary = ss[POSITION_DK_SALARY_INDEX]
                                ss_avg_score = ss[POSITION_DK_AVG_SCORE_INDEX]
                            else:
                                continue

                            for tb in thirdbase:

                                total_avg_score = p1_avg_score + p2_avg_score + c_avg_score + fb_avg_score \
                                                  + sb_avg_score + ss_avg_score
                                if total_avg_score + max_tb_avg_score + max_of_avg_score * 3 < MIN_SCORE_TARGET:
                                    continue

                                total_salary = p1_salary + p2_salary + c_salary + fb_salary + sb_salary + ss_salary
                                if total_salary + max_tb_salary + max_of_salary * 3 < MIN_SALARY_TARGET \
                                        or total_salary + min_tb_salary + min_of_salary * 3 > MAX_SALARY_TARGET:
                                    continue

                                tb_id = tb[POSITION_DK_ID_INDEX]
                                if c_id != tb_id and fb_id != tb_id and sb_id != tb_id and ss_id != tb_id:
                                    tb_teamid = tb[POSITION_TEAMID_INDEX]
                                    tb_batting_order = tb[POSITION_BATTING_ORDER_INDEX]
                                    tb_player_team = tb[POSITION_PLAYER_TEAM_INDEX]
                                    tb_salary = tb[POSITION_DK_SALARY_INDEX]
                                    tb_avg_score = tb[POSITION_DK_AVG_SCORE_INDEX]
                                else:
                                    continue

                                if c_teamid == fb_teamid == sb_teamid == ss_teamid == tb_teamid \
                                        and (p1_teamid == c_teamid or p2_teamid == c_teamid):
                                    continue

                                for of1 in range(0, len(outfield) - 2):

                                    total_avg_score = p1_avg_score + p2_avg_score + c_avg_score + fb_avg_score \
                                                      + sb_avg_score + ss_avg_score + tb_avg_score
                                    if total_avg_score + max_of_avg_score * 3 < MIN_SCORE_TARGET:
                                        continue

                                    total_salary = p1_salary + p2_salary + c_salary + fb_salary + sb_salary \
                                                    + ss_salary + tb_salary
                                    if total_salary + max_of_salary * 3 < MIN_SALARY_TARGET \
                                            or total_salary + min_of_salary * 3 > MAX_SALARY_TARGET:
                                        continue

                                    of1_id = outfield[of1][POSITION_DK_ID_INDEX]
                                    if c_id != of1_id and fb_id != of1_id and sb_id != of1_id and ss_id != of1_id \
                                            and tb_id != of1_id:
                                        of1_teamid = outfield[of1][POSITION_TEAMID_INDEX]
                                        of1_batting_order = outfield[of1][POSITION_BATTING_ORDER_INDEX]
                                        of1_player_team = outfield[of1][POSITION_PLAYER_TEAM_INDEX]
                                        of1_salary = outfield[of1][POSITION_DK_SALARY_INDEX]
                                        of1_avg_score = outfield[of1][POSITION_DK_AVG_SCORE_INDEX]
                                    else:
                                        continue

                                    for of2 in range(of1 + 1, len(outfield) - 1):

                                        total_avg_score = p1_avg_score + p2_avg_score + c_avg_score + fb_avg_score \
                                                          + sb_avg_score + ss_avg_score + tb_avg_score + of1_avg_score
                                        if total_avg_score + max_of_avg_score * 2 < MIN_SCORE_TARGET:
                                            continue

                                        total_salary = p1_salary + p2_salary + c_salary + fb_salary + sb_salary \
                                                       + ss_salary + tb_salary + of1_salary
                                        if total_salary + max_of_salary * 2 < MIN_SALARY_TARGET \
                                                or total_salary + min_of_salary * 2 > MAX_SALARY_TARGET:
                                            continue

                                        of2_id = outfield[of2][POSITION_DK_ID_INDEX]
                                        if c_id != of2_id and fb_id != of2_id and sb_id != of2_id and ss_id != of2_id \
                                                and tb_id != of2_id:
                                            of2_teamid = outfield[of2][POSITION_TEAMID_INDEX]
                                            of2_batting_order = outfield[of2][POSITION_BATTING_ORDER_INDEX]
                                            of2_player_team = outfield[of2][POSITION_PLAYER_TEAM_INDEX]
                                            of2_salary = outfield[of2][POSITION_DK_SALARY_INDEX]
                                            of2_avg_score = outfield[of2][POSITION_DK_AVG_SCORE_INDEX]
                                        else:
                                            continue

                                        for of3 in range(of2 + 1, len(outfield)):

                                            total_avg_score = p1_avg_score + p2_avg_score + c_avg_score + fb_avg_score \
                                                              + sb_avg_score + ss_avg_score + tb_avg_score \
                                                              + of1_avg_score + of2_avg_score
                                            if total_avg_score + max_of_avg_score < MIN_SCORE_TARGET:
                                                continue

                                            total_salary = p1_salary + p2_salary + c_salary + fb_salary + sb_salary \
                                                           + ss_salary + tb_salary + of1_salary + of2_salary
                                            if total_salary + max_of_salary < MIN_SALARY_TARGET \
                                                    or total_salary + min_of_salary > MAX_SALARY_TARGET:
                                                continue

                                            of3_id = outfield[of3][POSITION_DK_ID_INDEX]
                                            if c_id != of3_id and fb_id != of3_id and sb_id != of3_id \
                                                    and ss_id != of3_id and tb_id != of3_id:
                                                of3_teamid = outfield[of3][POSITION_TEAMID_INDEX]
                                                of3_batting_order = outfield[of3][POSITION_BATTING_ORDER_INDEX]
                                                of3_player_team = outfield[of3][POSITION_PLAYER_TEAM_INDEX]
                                                of3_salary = outfield[of3][POSITION_DK_SALARY_INDEX]
                                                of3_avg_score = outfield[of3][POSITION_DK_AVG_SCORE_INDEX]
                                            else:
                                                continue

                                            #teamids = [p1_teamid, p2_teamid, c_teamid, fb_teamid, sb_teamid, ss_teamid,
                                            #           tb_teamid, of1_teamid, of2_teamid, of3_teamid]
                                            max_freq_not_met = True
                                            #for i in teamids:
                                           #    freq = teamids.count(i)
                                            #    if freq > MAX_PLAYERS_PER_TEAM:
                                            #        max_freq_not_met = False
                                            #        break

                                            total_salary = p1_salary + p2_salary + c_salary + fb_salary + sb_salary \
                                                           + ss_salary + tb_salary + of1_salary + of2_salary \
                                                           + of3_salary
                                            total_avg_score = p1_avg_score + p2_avg_score + c_avg_score + fb_avg_score \
                                                              + sb_avg_score + ss_avg_score + tb_avg_score \
                                                              + of1_avg_score + of2_avg_score + of3_avg_score
                                            if MAX_SALARY_TARGET >= total_salary >= MIN_SALARY_TARGET \
                                                    and total_avg_score > MIN_SCORE_TARGET and max_freq_not_met:
                                                teams.append([p1_teamid, p1_id, p2_teamid, p2_id,
                                                              c_teamid, c_batting_order, c_id, fb_teamid,
                                                              fb_batting_order, fb_id, sb_teamid,
                                                              sb_batting_order, sb_id, tb_teamid,
                                                              tb_batting_order, tb_id, ss_teamid,
                                                              ss_batting_order, ss_id, of1_teamid,
                                                              of1_batting_order, of1_id, of2_teamid,
                                                              of2_batting_order, of2_id, of3_teamid,
                                                              of3_batting_order, of3_id, total_salary,
                                                              0, 0, 0, 0, 0])


def combineTeamsAndScores(hitterscores, pitcherscores, finaloutput):
    global start_time
    global outputfile
    global dk_teams

    for gameid in range(1, GAMES_TO_SIM + 1):
        for team in dk_teams:
            p1_index = getPitcherStatsIndex(gameid, team[P1_TEAMID_INDEX], 1)
            p2_index = getPitcherStatsIndex(gameid, team[P2_TEAMID_INDEX], 1)
            c_index = getStatsIndex(gameid, team[C_TEAMID_INDEX], team[C_BATTING_ORDER_INDEX])
            fb_index = getStatsIndex(gameid, team[FB_TEAMID_INDEX], team[FB_BATTING_ORDER_INDEX])
            sb_index = getStatsIndex(gameid, team[SB_TEAMID_INDEX], team[SB_BATTING_ORDER_INDEX])
            ss_index = getStatsIndex(gameid, team[SS_TEAMID_INDEX], team[SS_BATTING_ORDER_INDEX])
            tb_index = getStatsIndex(gameid, team[TB_TEAMID_INDEX], team[TB_BATTING_ORDER_INDEX])
            of1_index = getStatsIndex(gameid, team[OF1_TEAMID_INDEX], team[OF1_BATTING_ORDER_INDEX])
            of2_index = getStatsIndex(gameid, team[OF2_TEAMID_INDEX], team[OF2_BATTING_ORDER_INDEX])
            of3_index = getStatsIndex(gameid, team[OF3_TEAMID_INDEX], team[OF3_BATTING_ORDER_INDEX])
            total_score = pitcherscores[p1_index][P_SCORE_INDEX] + pitcherscores[p2_index][P_SCORE_INDEX] \
                          + hitterscores[c_index][B_SCORE_INDEX] + hitterscores[fb_index][B_SCORE_INDEX] \
                          + hitterscores[sb_index][B_SCORE_INDEX] + hitterscores[ss_index][B_SCORE_INDEX] \
                          + hitterscores[tb_index][B_SCORE_INDEX] + hitterscores[of1_index][B_SCORE_INDEX] \
                          + hitterscores[of2_index][B_SCORE_INDEX] + hitterscores[of3_index][B_SCORE_INDEX]
#            finaloutput.append([gameid, team[P1_PLAYER_TEAM_INDEX], team[P2_PLAYER_TEAM_INDEX],
#                                team[C_PLAYER_TEAM_INDEX], team[FB_PLAYER_TEAM_INDEX], team[SB_PLAYER_TEAM_INDEX],
#                                team[SS_PLAYER_TEAM_INDEX], team[TB_PLAYER_TEAM_INDEX], team[OF1_PLAYER_TEAM_INDEX],
#                                team[OF2_PLAYER_TEAM_INDEX], team[OF3_PLAYER_TEAM_INDEX], total_score,
#                                team[TOTAL_SALARY_INDEX]])

            team[TEAM_TOTAL_SCORE_INDEX] = team[TEAM_TOTAL_SCORE_INDEX] + total_score

            if total_score >= SCORE_TARGET_4:
                team[SCORE_TARGET_4_COUNT_INDEX] += 1
                team[SCORE_TARGET_3_COUNT_INDEX] += 1
                team[SCORE_TARGET_2_COUNT_INDEX] += 1
                team[SCORE_TARGET_1_COUNT_INDEX] += 1
            if total_score >= SCORE_TARGET_3:
                team[SCORE_TARGET_3_COUNT_INDEX] += 1
                team[SCORE_TARGET_2_COUNT_INDEX] += 1
                team[SCORE_TARGET_1_COUNT_INDEX] += 1
            if total_score >= SCORE_TARGET_2:
                team[SCORE_TARGET_2_COUNT_INDEX] += 1
                team[SCORE_TARGET_1_COUNT_INDEX] += 1
            if total_score >= SCORE_TARGET_1:
                team[SCORE_TARGET_1_COUNT_INDEX] += 1

        if gameid % 100 == 0:
            print("%f minutes - game " % ((time.time() - start_time) / 60) + str(gameid) + " dk scores recorded")


    for team in dk_teams:
        outputfile.write("\n")
        outputfile.write(str(team[P1_PLAYER_TEAM_INDEX]) + ", "
                         + str(team[P2_PLAYER_TEAM_INDEX]) + ", " + str(team[C_PLAYER_TEAM_INDEX]) + ", "
                         + str(team[FB_PLAYER_TEAM_INDEX]) + ", " + str(team[SB_PLAYER_TEAM_INDEX]) + ", "
                         + str(team[SS_PLAYER_TEAM_INDEX]) + ", " + str(team[TB_PLAYER_TEAM_INDEX]) + ", "
                         + str(team[OF1_PLAYER_TEAM_INDEX]) + ", " + str(team[OF2_PLAYER_TEAM_INDEX]) + ", "
                         + str(team[OF3_PLAYER_TEAM_INDEX]) + ", " + str(team[TEAM_TOTAL_SCORE_INDEX]) + ", "
                         + str(team[TOTAL_SALARY_INDEX]) + ", " + str(team[SCORE_TARGET_1_COUNT_INDEX]) + ", "
                         + str(team[SCORE_TARGET_2_COUNT_INDEX]) + ", " + str(team[SCORE_TARGET_3_COUNT_INDEX]) + ", "
                         + str(team[SCORE_TARGET_4_COUNT_INDEX]))

start_time = time.time()

battersfile = open("outputbatters.csv",'w')
pitchersfile = open("outputpitchers.csv",'w')
teamsfile = open("outputteams.csv",'w')
testfile = open("output.csv",'w')

path ='c:/Users/mmax/Documents/BasedballInput.xlsx'
workbook = openpyxl.load_workbook(path)
batterssheet = workbook["BattersForInput"]
batters = []
lastteam = getBatters(batterssheet, batters)
battercount = len(batters)
pitcherssheet = workbook["PitchersForInput"]
pitchers = []
getPitchers(pitcherssheet, pitchers)
pitchercount = len(pitchers)

print("%f minutes - files are loaded" % ((time.time() - start_time)/60))

# initiate stats arrays
batterstats = []
pitcherstats = []
teamstats = []

for game in range(1, GAMES_TO_SIM + 1):
    for batter in batters:
        batterstats.append([batter[5], game, batter[6], batter[0], batter[1], batter[2], batter[3], batter[4],
                                0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

    for pitcher in pitchers:
        pitcherstats.append([pitcher[0], game, pitcher[1], pitcher[2], pitcher[3], pitcher[4], 0, 0, 0, 0, 0, 0,
                             0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

    for team in range(1, lastteam + 1):
        teamstats.append([batters[9*(team-1)][5], game, batters[9*(team-1)][2], batters[9*(team-1)][3], 0, 0])

print("%f minutes - stat sheets are initiated" % ((time.time() - start_time)/60))

# set game to beginning defaults
team1batter = 1
team2batter = 1
inning = 1
team1 = 1
team2 = 2
team1runs = 0
team2runs = 0
win = False
winningteam = 0
first = [0,0,0]
second = [0,0,0]
third = [0,0,0]
outs = 0
team1pitcher = 1
team2pitcher = 1

print("%f minutes - games have begun" % ((time.time() - start_time)/60))

while team2 <= lastteam:
    for gameid in range(1, GAMES_TO_SIM + 1):

        team1runs = 0
        team2runs = 0
        team1batter = 1
        team2batter = 1
        inning = 1
        win = False
        winningteam = 0
        losingteam = 0
        outs = 0
        first = [0, 0, 0]
        second = [0, 0, 0]
        third = [0, 0, 0]
        team1pitcher = 1
        team2pitcher = 1


        while win == False:

            # run away team inning
            while outs < 3:

                currentbatter = team1batter
                team2pitcher = checkforpitchingchange(gameid, team2, team2pitcher, team2runs, team1runs)
                calculateAtBat(gameid, team1, team1batter, team2, team2pitcher)
                if team1batter == 9:
                    team1batter = 1
                else:
                    team1batter = team1batter + 1

            if inning == 9:
                if team1runs < team2runs:
                    win = True
                    winningteam = team2
                    losingteam = team1


            # clear the basses and outs to start the next half inning
            outs = 0
            first = [0, 0, 0]
            second = [0, 0, 0]
            third = [0, 0, 0]

            # run home team inning
            while outs < 3 and win == False:

                currentbatter = team2batter
                team1pitcher = checkforpitchingchange(gameid, team1, team1pitcher, team1runs, team2runs)
                calculateAtBat(gameid, team2, team2batter, team1, team1pitcher)
                if team2batter == 9:
                    team2batter = 1
                else:
                    team2batter = team2batter + 1

            if inning >= 9:
                if team1runs < team2runs:
                    win = True
                    winningteam = team2
                    losingteam = team1
                elif team2runs < team1runs:
                    win = True
                    winningteam = team1
                    losingteam = team2

            # clear the basses and outs to start the next half inning
            outs = 0
            first = [0, 0, 0]
            second = [0, 0, 0]
            third = [0, 0, 0]
            inning = inning + 1

        # update teams sheet with winner and score
        teamstats[getTeamStatsIndex(gameid, team1)][5] = team1runs
        teamstats[getTeamStatsIndex(gameid, team2)][5] = team2runs
        teamstats[getTeamStatsIndex(gameid, winningteam)][4] = 1
        team1pitcherindex = getPitcherStatsIndex(gameid, team1, SP_INDEX)
        team2pitcherindex = getPitcherStatsIndex(gameid, team2, SP_INDEX)
        team1relieverindex = getPitcherStatsIndex(gameid, team1, RP_INDEX)
        team2relieverindex = getPitcherStatsIndex(gameid, team2, RP_INDEX)
        pitcherstats[team1pitcherindex][P_TRAE_INDEX] = team1runs
        pitcherstats[team2pitcherindex][P_TRAE_INDEX] = team2runs
        pitcherstats[team1pitcherindex][P_ORAE_INDEX] = team2runs
        pitcherstats[team2pitcherindex][P_ORAE_INDEX] = team1runs
        team1relieverouts = pitcherstats[team1relieverindex][P_O_INDEX]
        team2relieverouts = pitcherstats[team2relieverindex][P_O_INDEX]
        if teamstats[getTeamStatsIndex(gameid, losingteam)][5] \
                < pitcherstats[getPitcherStatsIndex(gameid, winningteam, SP_INDEX)][P_TRAT_INDEX]:
            pitcherstats[getPitcherStatsIndex(gameid, winningteam, SP_INDEX)][P_W_INDEX] = 1
        if team1relieverouts == 0:
            pitcherstats[team1pitcherindex][P_CG_INDEX] = 1
            if pitcherstats[team1pitcherindex][P_ORAE_INDEX] == 0:
                pitcherstats[team1pitcherindex][P_CGS_INDEX] = 1
            if pitcherstats[team1pitcherindex][P_H_INDEX] == 0:
                pitcherstats[team1pitcherindex][P_NH_INDEX] = 1
        if team2relieverouts == 0:
            pitcherstats[team2pitcherindex][P_CG_INDEX] = 1
            if pitcherstats[team2pitcherindex][P_ORAE_INDEX] == 0:
                pitcherstats[team2pitcherindex][P_CGS_INDEX] = 1
            if pitcherstats[team2pitcherindex][P_H_INDEX] == 0:
                pitcherstats[team2pitcherindex][P_NH_INDEX] = 1

    print("%f minutes - " % ((time.time() - start_time)/60) + str(team1) + " vs. " + str(team2) + " games done")

    team1 = team1 + 2
    team2 = team2 + 2

print("%f minutes - games have ended" % ((time.time() - start_time)/60))

outputfile = open("scoresoutput.csv", 'w')

battersfile.write("teamid, gameid, battingorder, playerid, playername, teamname, gamename, pos, PA, H, 1B, 2B, 3B, "
                  "HR, O, SO, FO, GO, BB, RBI, R, SB, CS, Score")
for line in batterstats:
    line[B_SCORE_INDEX] = line[SINGLE_INDEX] * 3 + line[DOUBLE_INDEX] * 5 + line[TRIPLE_INDEX] * 8 + line[HR_INDEX] * 10 \
                   + line[RBI_INDEX] * 2 + line[R_INDEX] * 2 + line[BB_INDEX] * 2 + line[SB_INDEX] * 5 \
                   - line[CS_INDEX] * 2
    battersfile.write("\n")
    for item in line:
        battersfile.write(str(item) + ", ")

pitchersfile.write("teamid, gameid, pitchingorder, playername, teamname, gamename, TBF, H, 1B, 2B, 3B, HR, BB, "
                  "SO, O, ER, teamrunsatexit, opprunsatexit, teamsrunsatend, opprunsatend, w, cg, csg, nh, score")
for line in pitcherstats:
    line[P_SCORE_INDEX] = line[P_O_INDEX] * .75 + line[P_SO_INDEX] * 2 + line[P_W_INDEX] * 4 - line[P_ER_INDEX] * 2 \
                          - line[P_H_INDEX] * .6 - line[P_BB_INDEX] * .6 + line[P_CG_INDEX] * 2.5 \
                          + line[P_CGS_INDEX] * 2.5 + line[P_NH_INDEX] * 5
    pitchersfile.write("\n")
    for item in line:
        pitchersfile.write(str(item) + ", ")

teamsfile.write("teamid, gameid, teamname, gamename, w, r")
for line in teamstats:
    teamsfile.write("\n")
    for item in line:
        teamsfile.write(str(item) + ", ")

print("%f minutes - game stats have been recorded" % ((time.time() - start_time)/60))

# Draftkings stuff is down here
catchers = []
firstbase = []
secondbase = []
shortstop = []
thirdbase = []
outfield = []
dk_teams = []
finaloutput = []
max_salary = 0
min_salary = 50000
max_ss_salary = 0
min_ss_salary = 50000
max_tb_salary = 0
min_tb_salary = 50000
max_of_salary = 0
min_of_salary = 50000
max_c_avg_score = 0
max_fb_avg_score = 0
max_sb_avg_score = 0
max_ss_avg_score = 0
max_tb_avg_score = 0
max_of_avg_score = 0

getHitters(batterssheet, catchers, firstbase, secondbase, shortstop, thirdbase, outfield)

print("%f minutes - hitters have been separated by position" % ((time.time() - start_time)/60))

initiateTeams(dk_teams, pitchers, catchers, firstbase, secondbase, shortstop, thirdbase, outfield)

print("%f minutes - " % ((time.time() - start_time)/60) + str(len(dk_teams)) + " dk teams have been created")

outputfile.write("P1, P2, C, 1B, 2B, 3B, SS, OF1, OF2, OF3, total scores, total salary, score target 1 count, "
                 "score target 2 count, score target 3 count, score target 4 count")
combineTeamsAndScores(batterstats, pitcherstats, finaloutput)

print("%f minutes - dk teams and game stats have been combined" % ((time.time() - start_time)/60))


#for line in finaloutput:
#    outputfile.write("\n")
#    for item in line:
#        outputfile.write(str(item) + ", ")

print("%f minutes - dk teams and scores have been recorded" % ((time.time() - start_time)/60))

#testfile.write("teamid, pitcherorder, playername, teamname, gamename, tbf_target, er_target, single_coef, "
#               "double_coef, triple_coef, hr_coef, bb_coef, so_coef, dk id, dk salary")
#for line in pitchers:
#    testfile.write("\n")
#    for item in line:
#        testfile.write(str(item) + ", ")

#testfile.write("playerid, playername, teamname, gamename, pos, team, battingorder, single, double, triple, HR, "
#                       "BB, SO, FO, GO, SB, CS, dk_id, dk_pos, dk_salary")
#for line in batters:
#    testfile.write("\n")
#    for item in line:
#        testfile.write(str(item) + ", ")

