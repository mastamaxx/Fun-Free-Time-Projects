import openpyxl
import csv
import GlobalVariables
import time

GAMES_TO_SIM = GlobalVariables.GAMES_TO_SIM
MIN_SALARY_TARGET = GlobalVariables.MIN_SALARY_TARGET
MAX_SALARY_TARGET = GlobalVariables.MAX_SALARY_TARGET
PLAYERS_PER_DK_TEAM = GlobalVariables.PLAYERS_PER_DK_TEAM
MAX_PLAYERS_PER_TEAM = GlobalVariables.MAX_PLAYERS_PER_TEAM

PA_INDEX = GlobalVariables.PA_INDEX
H_INDEX = GlobalVariables.H_INDEX
SINGLE_INDEX = GlobalVariables.SINGLE_INDEX
DOUBLE_INDEX = GlobalVariables.DOUBLE_INDEX
TRIPLE_INDEX = GlobalVariables.TRIPLE_INDEX
HR_INDEX = GlobalVariables.HR_INDEX
O_INDEX = GlobalVariables.O_INDEX
SO_INDEX = GlobalVariables.SO_INDEX
FO_INDEX = GlobalVariables.FO_INDEX
GO_INDEX = GlobalVariables.GO_INDEX
BB_INDEX = GlobalVariables.BB_INDEX
RBI_INDEX = GlobalVariables.RBI_INDEX
R_INDEX = GlobalVariables.R_INDEX
SB_INDEX = GlobalVariables.SB_INDEX
CS_INDEX = GlobalVariables.CS_INDEX
B_SCORE_INDEX = GlobalVariables.B_SCORE_INDEX

P_TBF_INDEX = GlobalVariables.P_TBF_INDEX
P_H_INDEX = GlobalVariables.P_H_INDEX
P_SINGLE_INDEX = GlobalVariables.P_SINGLE_INDEX
P_DOUBLE_INDEX = GlobalVariables.P_DOUBLE_INDEX
P_TRIPLE_INDEX = GlobalVariables.P_TRIPLE_INDEX
P_HR_INDEX = GlobalVariables.P_HR_INDEX
P_BB_INDEX = GlobalVariables.P_BB_INDEX
P_SO_INDEX = GlobalVariables.P_SO_INDEX
P_O_INDEX = GlobalVariables.P_O_INDEX
P_ER_INDEX = GlobalVariables.P_ER_INDEX
P_TRAT_INDEX = GlobalVariables.P_TRAT_INDEX
P_ORAT_INDEX = GlobalVariables.P_ORAT_INDEX
P_TRAE_INDEX = GlobalVariables.P_TRAE_INDEX
P_ORAE_INDEX = GlobalVariables.P_ORAE_INDEX
P_W_INDEX = GlobalVariables.P_W_INDEX
P_CG_INDEX = GlobalVariables.P_CG_INDEX
P_CGS_INDEX = GlobalVariables.P_CGS_INDEX
P_NH_INDEX = GlobalVariables.P_NH_INDEX
P_SCORE_INDEX = GlobalVariables.P_SCORE_INDEX

PITCHERS_PER_TEAM = GlobalVariables.PITCHERS_PER_TEAM
SP_INDEX = GlobalVariables.SP_INDEX
RP_INDEX = GlobalVariables.RP_INDEX

TEAM_1_FOR_MATCH = 1

PLAYER_TEAM_INDEX = 0
PLAYER_ORDER_INDEX = 1
PLAYER_POS_INDEX = 2
PLAYER_DK_ID_INDEX = 3
PLAYER_DK_CPT_SALARY_INDEX = 4
PLAYER_DK_UTIL_SALARY_INDEX = 5

CPT_TEAMID_INDEX = 0
CPT_ORDER_INDEX = 1
CPT_POS_INDEX = 2
CPT_ID_INDEX = 3
UTIL1_TEAMID_INDEX = 4
UTIL1_ORDER_INDEX = 5
UTIL1_POS_INDEX = 6
UTIL1_ID_INDEX = 7
UTIL2_TEAMID_INDEX = 8
UTIL2_ORDER_INDEX = 9
UTIL2_POS_INDEX = 10
UTIL2_ID_INDEX = 11
UTIL3_TEAMID_INDEX = 12
UTIL3_ORDER_INDEX = 13
UTIL3_POS_INDEX = 14
UTIL3_ID_INDEX = 15
UTIL4_TEAMID_INDEX = 16
UTIL4_ORDER_INDEX = 17
UTIL4_POS_INDEX = 18
UTIL4_ID_INDEX = 19
UTIL5_TEAMID_INDEX = 20
UTIL5_ORDER_INDEX = 21
UTIL5_POS_INDEX = 22
UTIL5_ID_INDEX = 23
SHOWDOWN_TEAM_SALARY_INDEX = 24
SHOWDOWN_TEAM_SCORE_INDEX = 25

def loadBatterStats():

    with open('outputbatters.csv', newline='') as csvfile:
        batterstats = list(csv.reader(csvfile))

    return batterstats


def loadPitcherStats():

    with open('outputpitchers.csv', newline='') as csvfile:
        pitcherstats = list(csv.reader(csvfile))

    return pitcherstats


def getPlayersForShowdown(battersheet, pitchersheet, players, team1, team2):

    for row in range(2, battersheet.max_row + 1):

        teamid = battersheet['U' + str(row)].value

        if teamid == team1 or teamid == team2:
            battingorder = battersheet['V' + str(row)].value
            pos = 'h'
            dk_id = battersheet['AG' + str(row)].value
            dk_showdown_cpt_salary = battersheet['AJ' + str(row)].value
            dk_showdown_util_salary = battersheet['AK' + str(row)].value

            players.append([teamid, battingorder, pos, dk_id, dk_showdown_cpt_salary, dk_showdown_util_salary])

    for row in range(2, pitchersheet.max_row + 1):

        teamid = pitchersheet['A' + str(row)].value
        pitchingorder = pitchersheet['B' + str(row)].value

        if (teamid == team1 or teamid == team2) and pitchingorder == 1:
            pos = 'p'
            dk_id = pitchersheet['AB' + str(row)].value
            dk_showdown_cpt_salary = pitchersheet['AD' + str(row)].value
            dk_showdown_util_salary = pitchersheet['AE' + str(row)].value

            players.append([teamid, pitchingorder, pos, dk_id, dk_showdown_cpt_salary, dk_showdown_util_salary])


def initiateShowdownTeams(players, playercount):

    dk_showdown_teams = []

    for cpt in players:
        cpt_id = cpt[PLAYER_DK_ID_INDEX]
        cpt_teamid = cpt[PLAYER_TEAM_INDEX]
        cpt_order = cpt[PLAYER_ORDER_INDEX]
        cpt_pos = cpt[PLAYER_POS_INDEX]
        cpt_salary = cpt[PLAYER_DK_CPT_SALARY_INDEX]

        for util1 in range(0, playercount - 4):
            util1_id = players[util1][PLAYER_DK_ID_INDEX]

            if util1_id != cpt_id:
                util1_teamid = players[util1][PLAYER_TEAM_INDEX]
                util1_order = players[util1][PLAYER_ORDER_INDEX]
                util1_pos = players[util1][PLAYER_POS_INDEX]
                util1_salary = players[util1][PLAYER_DK_UTIL_SALARY_INDEX]
            else:
                continue

            for util2 in range(util1 + 1, playercount - 3):
                util2_id = players[util2][PLAYER_DK_ID_INDEX]

                if util2_id != cpt_id:
                    util2_teamid = players[util2][PLAYER_TEAM_INDEX]
                    util2_order = players[util2][PLAYER_ORDER_INDEX]
                    util2_pos = players[util2][PLAYER_POS_INDEX]
                    util2_salary = players[util2][PLAYER_DK_UTIL_SALARY_INDEX]
                else:
                    continue

                for util3 in range(util2 + 1, playercount - 2):
                    util3_id = players[util3][PLAYER_DK_ID_INDEX]

                    if util3_id != cpt_id:
                        util3_teamid = players[util3][PLAYER_TEAM_INDEX]
                        util3_order = players[util3][PLAYER_ORDER_INDEX]
                        util3_pos = players[util3][PLAYER_POS_INDEX]
                        util3_salary = players[util3][PLAYER_DK_UTIL_SALARY_INDEX]
                    else:
                        continue

                    for util4 in range(util3 + 1, playercount - 1):
                        util4_id = players[util4][PLAYER_DK_ID_INDEX]

                        if util4_id != cpt_id:
                            util4_teamid = players[util4][PLAYER_TEAM_INDEX]
                            util4_order = players[util4][PLAYER_ORDER_INDEX]
                            util4_pos = players[util4][PLAYER_POS_INDEX]
                            util4_salary = players[util4][PLAYER_DK_UTIL_SALARY_INDEX]
                        else:
                            continue

                        for util5 in range(util4 + 1, playercount):
                            util5_id = players[util5][PLAYER_DK_ID_INDEX]

                            if util5_id != cpt_id:
                                util5_teamid = players[util5][PLAYER_TEAM_INDEX]
                                util5_order = players[util5][PLAYER_ORDER_INDEX]
                                util5_pos = players[util5][PLAYER_POS_INDEX]
                                util5_salary = players[util5][PLAYER_DK_UTIL_SALARY_INDEX]
                            else:
                                continue

                            max_freq_not_met = True
                            if cpt_teamid == util1_teamid == util2_teamid == util3_teamid == util4_teamid \
                                    == util5_teamid:
                                max_freq_not_met = False

                            total_salary = cpt_salary + util1_salary + util2_salary + util3_salary + util4_salary \
                                           + util5_salary

                            if MAX_SALARY_TARGET >= total_salary >= MIN_SALARY_TARGET and max_freq_not_met:
                                dk_showdown_teams.append([cpt_teamid, cpt_order, cpt_pos, cpt_id, util1_teamid,
                                                          util1_order, util1_pos, util1_id, util2_teamid, util2_order,
                                                          util2_pos, util2_id, util3_teamid, util3_order, util3_pos,
                                                          util3_id, util4_teamid, util4_order, util4_pos, util4_id,
                                                          util5_teamid, util5_order, util5_pos, util5_id, total_salary,
                                                          0])

    return dk_showdown_teams



def combineShowdownTeamsAndScores(hitterscores, pitcherscores):
    global start_time
    global outputfile
    global dk_showdown_teams

    for gameid in range(1, GAMES_TO_SIM + 1):
        for team in dk_showdown_teams:
            cpt_pos = team[CPT_POS_INDEX]
            util1_pos = team[UTIL1_POS_INDEX]
            util2_pos = team[UTIL2_POS_INDEX]
            util3_pos = team[UTIL3_POS_INDEX]
            util4_pos = team[UTIL4_POS_INDEX]
            util5_pos = team[UTIL5_POS_INDEX]
            cpt_index = getStatsIndex(gameid, team[CPT_TEAMID_INDEX], team[CPT_ORDER_INDEX], cpt_pos)
            util1_index = getStatsIndex(gameid, team[UTIL1_TEAMID_INDEX], team[UTIL1_ORDER_INDEX], util1_pos)
            util2_index = getStatsIndex(gameid, team[UTIL2_TEAMID_INDEX], team[UTIL2_ORDER_INDEX], util2_pos)
            util3_index = getStatsIndex(gameid, team[UTIL3_TEAMID_INDEX], team[UTIL3_ORDER_INDEX], util3_pos)
            util4_index = getStatsIndex(gameid, team[UTIL4_TEAMID_INDEX], team[UTIL4_ORDER_INDEX], util4_pos)
            util5_index = getStatsIndex(gameid, team[UTIL5_TEAMID_INDEX], team[UTIL5_ORDER_INDEX], util5_pos)
            cpt_score = float(getPlayerScore(cpt_pos, cpt_index, pitcherscores, hitterscores)) * 1.5
            util1_score = float(getPlayerScore(util1_pos, util1_index, pitcherscores, hitterscores))
            util2_score = float(getPlayerScore(util2_pos, util2_index, pitcherscores, hitterscores))
            util3_score = float(getPlayerScore(util3_pos, util3_index, pitcherscores, hitterscores))
            util4_score = float(getPlayerScore(util4_pos, util4_index, pitcherscores, hitterscores))
            util5_score = float(getPlayerScore(util5_pos, util5_index, pitcherscores, hitterscores))
            total_score = float(cpt_score + util1_score + util2_score + util3_score + util4_score + util5_score)

            team[SHOWDOWN_TEAM_SCORE_INDEX] = team[SHOWDOWN_TEAM_SCORE_INDEX] + total_score

        if gameid % 100 == 0:
            print("%f minutes - game " % ((time.time() - start_time) / 60) + str(gameid) + " dk scores recorded")

    for team in dk_showdown_teams:
        outputfile.write("\n")
        outputfile.write(str(team[CPT_ID_INDEX]) + ", " + str(team[UTIL1_ID_INDEX]) + ", " + str(team[UTIL2_ID_INDEX])
                         + ", " + str(team[UTIL3_ID_INDEX]) + ", " + str(team[UTIL4_ID_INDEX]) + ", "
                         + str(team[UTIL5_ID_INDEX]) + ", " + str(team[SHOWDOWN_TEAM_SALARY_INDEX]) + ", "
                         + str(team[SHOWDOWN_TEAM_SCORE_INDEX]))


def getStatsIndex(gameid, teamid, playerorder, pos):
    return getPitcherStatsIndex(gameid, teamid, playerorder) if pos == "p" \
        else getBatterStatsIndex(gameid, teamid, playerorder)


def getPitcherStatsIndex(gameid, teamid, pitcherorder):
    return ((gameid - 1) * pitchercount + (teamid - 1) * PITCHERS_PER_TEAM + pitcherorder)


def getBatterStatsIndex(gameid, teamid, battingorder):
    return ((gameid - 1) * battercount + (teamid - 1) * 9 + battingorder)


def getPlayerScore(pos, player_index, pitcherstats, batterstats):
    return pitcherstats[player_index][P_SCORE_INDEX] if pos == "p" else batterstats[player_index][B_SCORE_INDEX]


start_time = time.time()

team1 = TEAM_1_FOR_MATCH
team2 = team1 + 1
path ='c:/Users/mmax/Documents/BasedballInput.xlsx'
workbook = openpyxl.load_workbook(path)
battersheet = workbook["BattersForInput"]
pitchersheet = workbook["PitchersForInput"]
print("%f minutes - files are loaded" % ((time.time() - start_time)/60))
players = []
battercount = battersheet.max_row - 1
pitchercount = pitchersheet.max_row - 1

batterstats = loadBatterStats()
pitcherstats = loadPitcherStats()
print("%f minutes - game stats are loaded" % ((time.time() - start_time)/60))

getPlayersForShowdown(battersheet, pitchersheet, players, team1, team2)
print("%f minutes - players are loaded" % ((time.time() - start_time)/60))
numplayers = len(players)

if numplayers != 20:
    quit("error: incorrect number of players loaded. Should be 20, actual was %i" %numplayers)

dk_showdown_teams = initiateShowdownTeams(players, numplayers)
print("%f minutes - " % ((time.time() - start_time)/60) + str(len(dk_showdown_teams)) + " dk teams have been created")

outputfile = open("showdownoutput.csv", 'w')
outputfile.write("CPT, UTIL1, UTIL2, UTIL3, UTIL4, UTIL5, total salary, total scores")
combineShowdownTeamsAndScores(batterstats, pitcherstats)
print("%f minutes - dk showdown teams and scores have been recorded" % ((time.time() - start_time)/60))
